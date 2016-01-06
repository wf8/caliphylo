
import csv
from openpyxl import load_workbook
from Bio import SeqIO

wb = load_workbook("Priority List-1.xlsx")
ws = wb.get_sheet_by_name("California_species_alignment")

#c = ws['A4']
#d = ws.cell(row = 4, column = 2)

gene_labels = ["18S","atpB","ndhF","trnL-trnF"]
gene_columns = [6,8,10,12]

def get_binomial(description):
    words = description.split("_")
    return words[0] + "_" + words[1]


def check_for_gene(gene, species):
    for record in SeqIO.parse(open("../" + gene + "_final.fasta", "rU"), "fasta"):
        binomial = get_binomial(record.id)
        if binomial == species:
            return 1
    return 0

row_num = 0
for row in ws.rows:
    genus = row[2].value
    species = row[3].value
    if genus != "Genus":
        binomial = genus + "_" + species
        i = 0
        for i in range(4):
            cell_value = check_for_gene(gene_labels[i], binomial)
            ws.cell(row = row_num, column = gene_columns[i]).value = cell_value
    row_num += 1

wb.save("Priority List-1_updated.xlsx")
