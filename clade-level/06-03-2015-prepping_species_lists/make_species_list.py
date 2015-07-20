

import csv
from openpyxl import load_workbook
#from openpyxl import get_sheet_by_name


wb = load_workbook("Priority List-1.xlsx")
ws = wb.get_sheet_by_name("California_species_alignment")

#c = ws['A4']
#d = ws.cell(row = 4, column = 2)

accepted_binomials = []
#i = 1
#while i < 5260:
#    genus = ws.cell(row = i, column = 2)
#    species = ws.cell(row = i, column = 3)
#    accepted_binomials.append(genus + "_" + species)
#    i += 1

i = 0
for row in ws.rows:
    genus = row[2].value
    species = row[3].value
    if genus != "Genus":
        i += 1
        accepted_binomials.append(genus + "_" + species)

print("Found " + str(i) + " accepted binomials...")

wb = load_workbook("Synonyms.xlsx")
ws = wb.get_sheet_by_name("synonyms")

csv_rows = []
for binomial in accepted_binomials:
    csv_row = binomial
    for row in ws.rows:
        if row[2].value == binomial:
            if row[1].value != binomial:
                csv_row += "," + row[1].value
            if row[0].value != binomial:
                csv_row += "," + row[0].value
    csv_rows.append(csv_row)

print("Writing CSV...")

with open("species.csv", "wb") as csvfile:
    csvwriter = csv.writer(csvfile)
    for row in csv_rows:
        csvwriter.writerow(row)

print("Done.")


